### Required for core functionality
import argparse
import os
import re
import ijson
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
from collections import defaultdict
import math
import csv

try:
    from tqdm import tqdm
    tqdm_available = True
except ImportError:
    tqdm_available = False
    print("Consider installing tqdm for progress bar: 'pip install tqdm'")

### For direct integration with Excel workbook
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.utils import column_index_from_string
    xl_available = True
except ImportError:
    xl_available = False   

### For debugging only
import pprint

### Global variable(s)
cc_snapshot_seconds = 1 # the size of concurrency snapshots in seconds
excel_sheet = "Data" # the name of the Excel sheet where the data goes


### Ingest the data file
def ingest_file(file_path):
    print("Reading data file...", end="", flush=True)
    scans = []
    tmp_field_names = []
    with open(file_path, 'rb') as file:
        # Extract field names from the @odata.context string
        context = ijson.items(file, '@odata.context')
        context_str = next(context)
        pattern = r"#Scans\((.*?)\)"
        match = re.search(pattern, context_str)
        if match:
            fields_str = match.group(1)
            tmp_field_names = [field.strip() for field in fields_str.split(',')]
            # Adjust the field names here, using tmp_field_names
            field_names = [field.replace('(LanguageName', '') if 'ScannedLanguages' in field else field for field in tmp_field_names]

        # Reset file pointer and extract scan items
        file.seek(0)
        for scan in ijson.items(file, 'value.item'):
            scans.append(scan)

    print("completed!")
    return field_names, scans


### Calcuate the difference between two timestamps in seconds
def calculate_time_difference(t1, t2):
    dt1 = parse_date(t1)
    dt2 = parse_date(t2)
    time_diff = (dt2 - dt1).total_seconds()
    
    return time_diff


### Output a data structure to the Excel 'Data' sheet starting at the indicated cell (e.g., J4)
def write_to_excel(data, start_col, start_row):
    try:
        # Convert start_col from letters to a numerical index
        start_col_index = column_index_from_string(start_col)
        wb_sheet = workbook[excel_sheet]
        
        for row_offset, row_data in enumerate(data, start=0):
            for col_offset, value in enumerate(row_data, start=0):
                # Calculate actual row and column indices
                row_idx = start_row + row_offset
                col_idx = start_col_index + col_offset
                wb_sheet.cell(row=row_idx, column=col_idx, value=value)
    except IOError as e:
        print(f"IOError when writing to Excel file: {e}")
        return
    except Exception as e:
        print(f"Unexpected error when writing to the Excel file: {e}")
        return


# Output a data structure to a csv file
def write_to_csv(header, data, filename):
    try:
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            writer.writerows(data)
    except IOError as e:
        print(f"IOError when writing to file: {e}")
    except Exception as e:
        print(f"Unexpected error when creating/writing to the CSV file: {e}")


# Write a single complete scan record to the full data csv
def write_scan_to_full_csv(field_names, scan, writer):
    try:
        # Build a row by extracting each field from the scan in the order of field_names
        row = []
        for field in field_names:
            if field == 'ScannedLanguages':
                # Special handling for ScannedLanguages field to convert list of dicts to comma-separated string
                languages = scan.get(field, [])
                language_str = ', '.join(lang['LanguageName'] for lang in languages)
                row.append(language_str)
            else:
                # For all other fields, use the value as-is
                row.append(scan.get(field, ""))
        # Write the constructed row to the CSV file
        writer.writerow(row)
    except IOError as e:
        print(f"IOError when writing to file: {e}")
    except Exception as e:
        print(f"Unexpected error when creating/writing to the CSV file: {e}")


### Process the scan data.
# One single function will be more efficient but start to get messy. Brace youreself.
def process_scans(scans, full_csv):

    ### Define (most) variables and data structures

    # Aggregate Metrics: Store high level metrics such as sums, averages, maximums, and totals
    # Note that sum, avg, and max are always associated with more granual metrics from other structures
    aggregate_metrics = {
        'COUNT_yes_scans': 0, # number of scans that fully ran
        'COUNT_no_scans': 0, # number of no-scans due to no code change
        'COUNT_missing_scans': 0, # number of scans with no recorded LOC; this is currently collected but unused
        'COUNT_scans': 0, # total of yes and no scans; excludes missing scans because those scans
        'COUNT_full_scans': 0, # number of full scans requested (includes no_scans)
        'COUNT_incremental_scans': 0, # number of incremental scans requested (includes no_scans)

        'SUM_loc': 0, # sum of the lines of code scanned in the data set
        'SUM_failed_loc': 0, # sum of the filed lines of code scanned in the data set
        'AVG_loc_scan': 0, # average number of lines of code per scan
        'AVG_failed_loc_scan': 0, # average number of failed lines of code per scan
        'AVG_loc_day': 0, # average number of lines of code per day
        'MAX_loc_scan': 0, # maximum number of lines of code per scan
        'MAX_failed_loc_scan': 0, # maximum number of failed lines of code per scan
        'MAX_loc_day': 0, # maximum number of lines of code per day
        
        'SUM_total_results': 0, # sum of total scan results
        'SUM_high_results': 0, # sum of high scan results
        'SUM_medium_results': 0, # sum of medium scan results
        'SUM_low_results': 0, # sum of low scan results
        'SUM_info_results': 0, # sum of info scan results
        'AVG_total_results': 0, # average number of total scan results
        'AVG_high_results': 0, # average number of high scan results
        'AVG_medium_results': 0, # average number of medium scan results
        'AVG_low_results': 0, # average number of low scan results
        'AVG_info_results': 0, # average number of info scan results
        'MAX_total_results': 0, # maximum number of total scan results
        'MAX_high_results': 0, # maximum number of high scan results
        'MAX_medium_results': 0, # maximum number of medium scan results
        'MAX_low_results': 0, # maximum number of low scan results
        'MAX_info_results': 0, # maximum number of info scan results

        'COUNT_high_results_scans': 0, # count of scans with high results
        'COUNT_medium_results_scans': 0, # count of scans with high results
        'COUNT_low_results_scans': 0, # count of scans with high results
        'COUNT_info_results_scans': 0, # count of scans with high results
        'COUNT_zero_results_scans': 0, # count of scans with high results

        'SUM_source_pulling_time': 0, # sum of total source pulling time in seconds
        'SUM_queue_time': 0, # sum of total queue time in seconds
        'SUM_engine_scan_time': 0, # sum of total engine scan time in seconds
        'SUM_total_scan_time': 0, # sum of total total scan time in seconds
        'AVG_source_pulling_time': 0, # average source pulling time in seconds
        'AVG_queue_time': 0, # average queue time in seconds
        'AVG_engine_scan_time': 0, # average engine scan time in seconds
        'AVG_total_scan_time': 0, # average total scan time in seconds; note this is not a sum of other times but specific data field that may exceed the sum
        'MAX_source_pulling_time': 0, # maximum source pulling time in seconds
        'MAX_queue_time': 0, # maximum queue time in seconds
        'MAX_engine_scan_time': 0, # maximum engine scan time in seconds
        'MAX_total_scan_time': 0, # maximum total scan time in seconds

        'COUNT_mon_scans': 0, # count of scans occuring on a Monday
        'COUNT_tue_scans': 0, # count of scans occuring on a Tuesday
        'COUNT_wed_scans': 0, # count of scans occuring on a Wednedsay
        'COUNT_thu_scans': 0, # count of scans occuring on a Thurdsay
        'COUNT_fri_scans': 0, # count of scans occuring on a Friday
        'COUNT_sat_scans': 0, # count of scans occuring on a Saturday
        'COUNT_sun_scans': 0, # count of scans occuring on a Sunday
        'COUNT_weekday_scans': 0, # count of scans occuring on a weekday
        'COUNT_weekend_scans': 0, # count of scans occuring on a weekend
        'MAX_scans_day': 0, # maximum number of scans per day
        'MAX_scan_date': None, # the date with the most scans

        'COUNT_projects_scanned': 0, # count of unique projects scanned

        'first_scan_date': datetime.max.date(), # the date of the first scan in the data set
        'last_scan_date': datetime.min.date(), # the date of the last scan in the data set
        'total_days': 0, # the number of days between the first and last scan
        'total_weeks': 0, # the number of weeks between the first and last scan 
        'total_scan_days': 0 # the totaly number of days that actually had scans
        }

    # Languages: Store language metrics (language_name, scan_count, scan_percentage) on a dynamic list of languages
    scan_languages = {}

    # Scan Origins: Store origin metrics; this is complicated because of many custom-named origins that need to be grouped
    scan_origins = {
    'ADO': {'printable_name': 'Azure DevOps', 'scan_count': 0, 'scan_percentage': 0},
    'Bamboo': {'printable_name': 'Bamboo', 'scan_count': 0, 'scan_percentage': 0},
    'CLI': {'printable_name': 'CLI', 'scan_count': 0, 'scan_percentage': 0},
    'cx-CLI': {'printable_name': 'CxCLI', 'scan_count': 0, 'scan_percentage': 0},
    'CxFlow': {'printable_name': 'CxFlow', 'scan_count': 0, 'scan_percentage': 0},
    'Eclipse': {'printable_name': 'Eclipse', 'scan_count': 0, 'scan_percentage': 0},
    'cx-intellij': {'printable_name': 'IntelliJ', 'scan_count': 0, 'scan_percentage': 0},
    'Jenkins': {'printable_name': 'Jenkins', 'scan_count': 0, 'scan_percentage': 0},
    'Manual': {'printable_name': 'Manual', 'scan_count': 0, 'scan_percentage': 0},
    'Maven': {'printable_name': 'Maven', 'scan_count': 0, 'scan_percentage': 0},
    'Other': {'printable_name': 'Other', 'scan_count': 0, 'scan_percentage': 0},
    'System': {'printable_name': 'System', 'scan_count': 0, 'scan_percentage': 0},
    'TeamCity': {'printable_name': 'TeamCIty', 'scan_count': 0, 'scan_percentage': 0},
    'TFS': {'printable_name': 'TFS', 'scan_count': 0, 'scan_percentage': 0},
    'Visual Studio': {'printable_name': 'Visual Studio', 'scan_count': 0, 'scan_percentage': 0},
    'Visual-Studio-Code': {'printable_name': 'VS Code', 'scan_count': 0, 'scan_percentage': 0},
    'VSTS': {'printable_name': 'VSTS', 'scan_count': 0, 'scan_percentage': 0},
    'Web Portal': {'printable_name': 'Web Portal', 'scan_count': 0, 'scan_percentage': 0}
    }

    # Scan Presets: Store preset metrics (preset_name, scan_count, scan_percentage) on a dynamic list of presets
    scan_presets = {}

    # Scan Times by LOC: Store various times for every scan grouped by LOC (source_pulling_time, queue_time, engine_scan_time, total_scan_time)
    scan_times_by_loc = {
        '0-20k': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '20k-50k': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '50k-100k': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '100k-250k': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '250k-500k': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '500k-1M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '1M-2M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '2M-3M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '3M-5M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '5M-7M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '7M-10M': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0},
        '10M+': {"COUNT_yes_scans": 0, "COUNT_no_scans": 0, "SUM_total_scan_time": 0, "SUM_source_pulling_time": 0, "SUM_queue_time": 0,
        "SUM_engine_scan_time": 0, "MAX_total_scan_time": 0, "MAX_source_pulling_time": 0, "MAX_queue_time": 0, "MAX_engine_scan_time": 0,
        "AVG_total_scan_time": 0, "AVG_source_pulling_time": 0, "AVG_queue_time": 0, "AVG_engine_scan_time": 0}
    }

    # Scan Statistics by Date: Store various statistics for every scan grouped by scan date
    scan_stats_by_date = {}

    # Temporary structure to track unique projects
    temp_pids = set()

    # Variables for concurrency
    # Event format: (timestamp, change_in_count, event_type)
    # Snapshot format: (timestamp, active_engines, queue_length)
    # change_in_count is +1 for starts (entering queue or starting engine) and -1 for ends (leaving queue or engine finishing)
    # event_type distinguishes between 'queue' and 'engine'
    cc_events = filtered_cc_events = snapshot_metrics = []
    
    ### Prepare to output CSV of all scan data and create output file, if required
    if full_csv['enabled']:
        try:
            filename = os.path.join(full_csv['csv_dir'], f"00-full_scan_data.csv")
            full_csv_file = open(filename, mode='w', newline='', encoding='utf-8')
            full_csv_writer = csv.writer(full_csv_file)
            full_csv_writer.writerow(full_csv['field_names'])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    ### Initialize tqdm object; we exclude concurrency processing because it's so fast, even for massive data sets
    if tqdm_available:
        pbar = tqdm(total=len(scans), desc="Processing scans")
    else:
        print("Processing scans...", end="", flush=True)

    ### Scan processing loop
    for scan in scans:
        if tqdm_available:
            pbar.update(1)
            pbar.refresh()
        
        # If required, we want to output to the full scan CSV first so as to include scans with missing fields (such as loc). This will cause a potential
        # mismatch between record counts but shouldn't impact anything relating to metrics or analysis. This CSV is only used for manual analysis.
        if full_csv['enabled']:
            write_scan_to_full_csv(full_csv['field_names'], scan, full_csv_writer)

        # If there is no LOC value, we might as well just completely skip the scan.
        # This differs from the current process but ensures that scan counts actually match in various metrics. 
        # We will record the missing scan.
        loc = scan.get('LOC', None)
        if loc is None:
            aggregate_metrics['COUNT_missing_scans'] += 1
            continue
        else:
            aggregate_metrics['COUNT_scans'] += 1

        scan_date_str = scan.get('ScanRequestedOn', '').split('T')[0]
        scan_date = datetime.strptime(scan_date_str, "%Y-%m-%d").date()

        ### Populate and update key aggregate metrics; note that averages have to be calculated later and many metrics are addressed later
        if scan.get('EngineFinishedOn', None) is not None:
            noscan = False
            aggregate_metrics['COUNT_yes_scans'] += 1
        else:
            noscan = True
            aggregate_metrics['COUNT_no_scans'] += 1

        if scan.get('IsIncremental', None):
            aggregate_metrics['COUNT_incremental_scans'] += 1
        else:
            aggregate_metrics['COUNT_full_scans'] += 1

        aggregate_metrics['SUM_loc'] += loc
        aggregate_metrics['SUM_failed_loc'] += loc
        aggregate_metrics['SUM_failed_loc'] += scan.get('FailedLOC', 0)
        aggregate_metrics['MAX_loc_scan'] = max(aggregate_metrics['MAX_loc_scan'], loc)
        aggregate_metrics['MAX_failed_loc_scan'] = max(aggregate_metrics['MAX_failed_loc_scan'], scan.get('FailedLOC', 0))

        aggregate_metrics['SUM_total_results'] += scan.get('TotalVulnerabilities', 0)
        aggregate_metrics['SUM_high_results'] += scan.get('High', 0)
        aggregate_metrics['SUM_medium_results'] += scan.get('Medium', 0)
        aggregate_metrics['SUM_low_results'] += scan.get('Low', 0)
        aggregate_metrics['SUM_info_results'] += scan.get('Info', 0)
        aggregate_metrics['MAX_total_results'] = max(aggregate_metrics['MAX_total_results'], scan.get('TotalVulnerabilities', 0))
        aggregate_metrics['MAX_high_results'] = max(aggregate_metrics['MAX_high_results'], scan.get('High', 0))
        aggregate_metrics['MAX_medium_results'] = max(aggregate_metrics['MAX_medium_results'], scan.get('Medium', 0))
        aggregate_metrics['MAX_low_results'] = max(aggregate_metrics['MAX_low_results'], scan.get('Low', 0))
        aggregate_metrics['MAX_info_results'] = max(aggregate_metrics['MAX_info_results'], scan.get('Info', 0))
        if scan.get('High', 0) > 0:
            aggregate_metrics['COUNT_high_results_scans'] += 1
        if scan.get('Medium', 0) > 0:
            aggregate_metrics['COUNT_medium_results_scans'] += 1
        if scan.get('Low', 0) > 0:
            aggregate_metrics['COUNT_low_results_scans'] += 1
        if scan.get('Info', 0) > 0:
            aggregate_metrics['COUNT_info_results_scans'] += 1
        if scan.get('TotalVulnerabilities', 0) == 0:
            aggregate_metrics['COUNT_zero_results_scans'] += 1

        # Update time metrics
        source_pulling_time = math.ceil(calculate_time_difference(scan.get('ScanRequestedOn'),scan.get('QueuedOn')))
        aggregate_metrics['SUM_source_pulling_time'] += source_pulling_time
        aggregate_metrics['MAX_source_pulling_time'] = max(aggregate_metrics['MAX_source_pulling_time'], source_pulling_time)
        
        queue_time = math.ceil(calculate_time_difference(scan.get('QueuedOn'),scan.get('EngineStartedOn')))
        aggregate_metrics['SUM_queue_time'] += queue_time
        aggregate_metrics['MAX_queue_time'] = max(aggregate_metrics['MAX_queue_time'], queue_time)
        
        if noscan is False:
            engine_scan_time = math.ceil(calculate_time_difference(scan.get('EngineStartedOn'),scan.get('EngineFinishedOn')))
            aggregate_metrics['SUM_engine_scan_time'] += engine_scan_time
            aggregate_metrics['MAX_engine_scan_time'] = max(aggregate_metrics['MAX_engine_scan_time'], scan.get('Low', 0))
        
        total_scan_time = math.ceil(calculate_time_difference(scan.get('ScanRequestedOn'),scan.get('ScanCompletedOn')))
        aggregate_metrics['SUM_total_scan_time'] += total_scan_time
        aggregate_metrics['MAX_total_scan_time'] = max(aggregate_metrics['MAX_total_scan_time'], total_scan_time)

        # Increment the proper day counters
        day_of_week = scan_date.strftime('%A')
        day_key_map = {
            'Monday': 'COUNT_mon_scans',
            'Tuesday': 'COUNT_tue_scans',
            'Wednesday': 'COUNT_wed_scans',
            'Thursday': 'COUNT_thu_scans',
            'Friday': 'COUNT_fri_scans',
            'Saturday': 'COUNT_sat_scans',
            'Sunday': 'COUNT_sun_scans',
        }
        if day_of_week in day_key_map:
            day_key = day_key_map[day_of_week]
            aggregate_metrics[day_key] += 1
        if day_of_week in ['Saturday', 'Sunday']:
            aggregate_metrics['COUNT_weekend_scans'] += 1
        else:
            aggregate_metrics['COUNT_weekday_scans'] += 1
        
        aggregate_metrics['first_scan_date'] = min(aggregate_metrics['first_scan_date'], scan_date)
        aggregate_metrics['last_scan_date'] = max(aggregate_metrics['last_scan_date'], scan_date)

        # Increment unique project count; sometimes the Id or Name is empty so create a new key type
        project_id = scan.get('ProjectId', 0)
        project_name = scan.get('ProjectName', "")
        pid = str(project_id) + "_" + project_name
        if pid not in temp_pids:
            temp_pids.add(pid)
            aggregate_metrics['COUNT_projects_scanned'] += 1

        ### Add scanned languages
        for language in scan.get('ScannedLanguages', []):
            lang_name = language.get('LanguageName')
            if lang_name and lang_name != "Common":
                scan_languages[lang_name] = scan_languages.get(lang_name, 0) + 1

        ### Add scan origin
        origin_key = scan.get('Origin', 'Other')
        group = next((key for key in scan_origins if origin_key.startswith(key)), 'Other')
        scan_origins[group]['scan_count'] += 1

        ### Add scan preset
        preset_name = scan.get('PresetName')
        scan_presets[preset_name] = scan_presets.get(preset_name, 0) + 1

        ### Add scan times
        if loc <= 20000:
            bin_key = '0-20k'
        elif loc <= 50000:
            bin_key = '20k-50k'
        elif loc <= 100000:
            bin_key = '50k-100k'
        elif loc <= 250000:
            bin_key = '100k-250k'
        elif loc <= 500000:
            bin_key = '250k-500k'
        elif loc <= 1000000:
            bin_key = '500k-1M'
        elif loc <= 2000000:
            bin_key = '1M-2M'
        elif loc <= 3000000:
            bin_key = '2M-3M'
        elif loc <= 5000000:
            bin_key = '3M-5M'
        elif loc <= 7000000:
            bin_key = '5M-7M'
        elif loc <= 10000000:
            bin_key = '7M-10M'
        else:
            bin_key = '10M+'
            
        bin = scan_times_by_loc[bin_key]

        bin['SUM_source_pulling_time'] += source_pulling_time
        bin['SUM_queue_time'] += queue_time
        bin['SUM_total_scan_time'] += total_scan_time
        bin['MAX_source_pulling_time'] = max(source_pulling_time, bin['MAX_source_pulling_time'])
        bin['MAX_queue_time'] = max(queue_time, bin['MAX_queue_time'])
        bin['MAX_total_scan_time'] = max(total_scan_time, bin['MAX_total_scan_time'])

        if noscan is False:
            bin['COUNT_yes_scans'] += 1
            bin['SUM_engine_scan_time'] += engine_scan_time
            bin['MAX_engine_scan_time'] = max(engine_scan_time, bin['MAX_engine_scan_time'])
        else:
            bin['COUNT_no_scans'] += 1

        ### Populate scan statistics by date
        if scan_date not in scan_stats_by_date:
            scan_stats_by_date[scan_date] = {
                'COUNT_yes_scans': 0,
                'COUNT_no_scans': 0,
                'COUNT_scans': 0,
                'COUNT_full_scans': 0,
                'COUNT_incremental_scans': 0,
                'SUM_loc': 0,
                'MAX_loc': 0,
                'SUM_failed_loc': 0,
                'MAX_failed_loc': 0
            }

        if noscan is False:
            scan_stats_by_date[scan_date]['COUNT_yes_scans'] += 1
        else:
            scan_stats_by_date[scan_date]['COUNT_no_scans'] += 1
                
        if scan.get('IsIncremental', None):
            scan_stats_by_date[scan_date]['COUNT_incremental_scans'] += 1
        else:
            scan_stats_by_date[scan_date]['COUNT_full_scans'] += 1

        scan_stats_by_date[scan_date]['COUNT_scans'] += 1
        scan_stats_by_date[scan_date]['SUM_loc'] += loc
        scan_stats_by_date[scan_date]['MAX_loc'] = max(loc, scan_stats_by_date[scan_date]['MAX_loc'])
        scan_stats_by_date[scan_date]['SUM_failed_loc'] += scan.get('FailedLOC', 0)
        scan_stats_by_date[scan_date]['MAX_failed_loc'] = max(scan.get('FailedLOC', 0), scan_stats_by_date[scan_date]['MAX_failed_loc'])
        
    # End of scan processing loop
        
    ### Calculate metrics that require the full data set
    aggregate_metrics['total_days'] = (aggregate_metrics['last_scan_date'] - aggregate_metrics['first_scan_date']).days
    aggregate_metrics['total_weeks'] = math.ceil(aggregate_metrics['total_days'] / 7)
    aggregate_metrics['total_scan_days'] = len(scan_stats_by_date)

    aggregate_metrics['AVG_loc_scan'] = math.ceil(aggregate_metrics['SUM_loc'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_failed_loc_scan'] = math.ceil(aggregate_metrics['SUM_failed_loc'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_loc_day'] = math.ceil(aggregate_metrics['SUM_loc'] / (aggregate_metrics['total_days']))

    for date, stats in scan_stats_by_date.items():
        aggregate_metrics['MAX_loc_day'] = max(aggregate_metrics['MAX_loc_day'], stats['SUM_loc'])
        
        if stats['COUNT_scans'] > aggregate_metrics['MAX_scans_day']:
            aggregate_metrics['MAX_scans_day'] = stats['COUNT_scans']
            aggregate_metrics['MAX_scan_date'] = date

    for bin_key, bin in scan_times_by_loc.items():
        if (bin['COUNT_yes_scans'] + bin['COUNT_no_scans']) > 0:
            bin['AVG_source_pulling_time'] = math.ceil(bin['SUM_source_pulling_time'] / (bin['COUNT_yes_scans'] + bin['COUNT_no_scans']))
            bin['AVG_queue_time'] = math.ceil(bin['SUM_queue_time'] / (bin['COUNT_yes_scans'] + bin['COUNT_no_scans']))
            bin['AVG_total_scan_time'] = math.ceil(bin['SUM_total_scan_time'] / (bin['COUNT_yes_scans'] + bin['COUNT_no_scans']))
        if bin['COUNT_yes_scans'] > 0:
            bin['AVG_engine_scan_time'] = math.ceil(bin['SUM_engine_scan_time'] / bin['COUNT_yes_scans'])
            bin['AVG_total_scan_time'] = math.ceil(bin['SUM_total_scan_time'] / bin['COUNT_yes_scans'])
    
    if aggregate_metrics['COUNT_scans'] > 0:
        aggregate_metrics['AVG_source_pulling_time'] = math.ceil(bin['SUM_source_pulling_time'] / aggregate_metrics['COUNT_scans'])
        aggregate_metrics['AVG_queue_time'] = math.ceil(aggregate_metrics['SUM_queue_time'] / aggregate_metrics['COUNT_scans'])
        aggregate_metrics['AVG_total_scan_time'] = math.ceil(aggregate_metrics['SUM_total_scan_time'] / aggregate_metrics['COUNT_scans'])
    if aggregate_metrics['COUNT_yes_scans'] > 0:
        aggregate_metrics['AVG_engine_scan_time'] = math.ceil(aggregate_metrics['SUM_engine_scan_time'] / aggregate_metrics['COUNT_yes_scans'])
        aggregate_metrics['AVG_total_scan_time'] = math.ceil(aggregate_metrics['SUM_total_scan_time'] / aggregate_metrics['COUNT_yes_scans'])

    aggregate_metrics['AVG_total_results'] = math.ceil(aggregate_metrics['SUM_total_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_high_results'] = round(aggregate_metrics['SUM_high_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_medium_results'] = round(aggregate_metrics['SUM_medium_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_low_results'] = round(aggregate_metrics['SUM_low_results'] / aggregate_metrics['COUNT_scans'])
    aggregate_metrics['AVG_info_results']= round(aggregate_metrics['SUM_info_results'] / aggregate_metrics['COUNT_scans'])



    for origin, data in scan_origins.items():
        data['scan_percentage'] = math.ceil((data['scan_count'] / aggregate_metrics['COUNT_scans']) * 100)
    
    if tqdm_available:
        pbar.close()
    else:
        print("completed!")

    # Process concurrency events
    print("Calculating scan concurrency...", end="", flush=True)
   
    # Initialize variables
    cc_window_start_ts = datetime.combine(aggregate_metrics['first_scan_date'], datetime.min.time()).timestamp()
    cc_window_end_ts = datetime.combine(aggregate_metrics['last_scan_date'], datetime.min.time()).timestamp()
    num_snapshots = math.ceil((cc_window_end_ts - cc_window_start_ts) / cc_snapshot_seconds)

    # Filter out objects based on the window and sort them
    filtered_cc_events = [event for event in cc_events if cc_window_start_ts <= event[0] <= cc_window_end_ts]
    filtered_cc_events.sort(key=lambda x: x[0])

    current_active_engines = 0
    current_queue_length = 0
    event_index = 0
    snapshot_metrics = []
    
    # For each snapshot...
    for snapshot in range(num_snapshots):
        # Calculate the bounds of the snapshot in timestamp format
        snapshot_start_ts = cc_window_start_ts + snapshot * cc_snapshot_seconds
        next_snapshot_start_ts = snapshot_start_ts + cc_snapshot_seconds
        
        while event_index < len(filtered_cc_events) and filtered_cc_events[event_index][0] < next_snapshot_start_ts:
            event_time, change, event_type = filtered_cc_events[event_index]
            
            if event_type == 'engine':
                current_active_engines += change
            elif event_type == 'queue':
                current_queue_length += change
            
            event_index += 1
        
        # Convert snapshot_start_ts to datetime for recording
        snapshot_start_dt = datetime.fromtimestamp(snapshot_start_ts)

        # Append the metrics for the current snapshot to the list
        snapshot_metrics.append((snapshot_start_dt, current_active_engines, current_queue_length))
    print("completed!")

    # Close the CSV file if it's open
    if full_csv['enabled']:
        full_csv_file.close()

    return {
        'aggregate_metrics': aggregate_metrics,
        'scan_languages': scan_languages,
        'scan_origins': scan_origins,
        'scan_presets': scan_presets,
        'scan_times_by_loc': scan_times_by_loc,
        'scan_stats_by_date': scan_stats_by_date,
        'cc_metrics': snapshot_metrics
    }


### Output to the screen as well as create very specific CSVs.
# One single function for simplification / variable reuse (but a bit messy, as well).
def output_analysis(data, csv_config, excel_config):
    # Crunch a few numbers that are needed in a specific format
    daily_scan_counts = {}
    weekly_scan_counts = {}

    for scan_date, stats in data['scan_stats_by_date'].items():
        daily_scan_counts[scan_date] = stats['COUNT_scans']
        
        # Calculate the Monday of the current week
        monday_of_week = scan_date - timedelta(days=scan_date.weekday())
        
        # Add the count for the current week
        if monday_of_week not in weekly_scan_counts:
            weekly_scan_counts[monday_of_week] = stats['COUNT_scans']
        else:
            weekly_scan_counts[monday_of_week] += stats['COUNT_scans']
    
    # Identify daily max concurrency values based on the granular calculations made previously
    daily_maxima = defaultdict(lambda: {'actual': 0, 'optimal': 0})
    overall_max_actual = 0
    overall_max_optimal = 0
    overall_max_actual_dates = set()
    overall_max_optimal_dates = set()

    for snapshot in data['cc_metrics']:
        snapshot_dt, active_engines, queue_length = snapshot
        snapshot_date = snapshot_dt.date()
        optimal_concurrency = active_engines + queue_length

        # Update daily maximums
        daily_record = daily_maxima[snapshot_date]
        daily_record['actual'] = max(daily_record['actual'], active_engines)
        daily_record['optimal'] = max(daily_record['optimal'], optimal_concurrency)

        # Update overall maximums and their dates
        if daily_record['actual'] > overall_max_actual:
            overall_max_actual = daily_record['actual']
            overall_max_actual_dates = {snapshot_date}
        elif daily_record['actual'] == overall_max_actual:
            overall_max_actual_dates.add(snapshot_date)

        if daily_record['optimal'] > overall_max_optimal:
            overall_max_optimal = daily_record['optimal']
            overall_max_optimal_dates = {snapshot_date}
        elif daily_record['optimal'] == overall_max_optimal:
            overall_max_optimal_dates.add(snapshot_date)

    # Print Summary of Scans
    print(f"\nSummary of Scans ({data['aggregate_metrics']['first_scan_date']} to {data['aggregate_metrics']['last_scan_date']})")
    print("-" * 50)
    print(f"Total number of scans: {format(data['aggregate_metrics']['COUNT_scans'], ',')}")
    print(f"- Full Scans: {format(data['aggregate_metrics']['COUNT_full_scans'], ',')} ({(data['aggregate_metrics']['COUNT_full_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Incremental Scans: {format(data['aggregate_metrics']['COUNT_incremental_scans'])} ({(data['aggregate_metrics']['COUNT_incremental_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- No Code Change Scans: {format(data['aggregate_metrics']['COUNT_no_scans'], ',')} ({(data['aggregate_metrics']['COUNT_no_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Scans with High Results: {format(data['aggregate_metrics']['COUNT_high_results_scans'], ',')} ({(data['aggregate_metrics']['COUNT_high_results_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Scans with Medium Results: {format(data['aggregate_metrics']['COUNT_medium_results_scans'], ',')} ({(data['aggregate_metrics']['COUNT_medium_results_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Scans with Low Results: {format(data['aggregate_metrics']['COUNT_low_results_scans'], ',')} ({(data['aggregate_metrics']['COUNT_low_results_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Scans with Informational Results: {format(data['aggregate_metrics']['COUNT_info_results_scans'], ',')} ({(data['aggregate_metrics']['COUNT_info_results_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Scans with Zero Results: {format(data['aggregate_metrics']['COUNT_zero_results_scans'], ',')} ({(data['aggregate_metrics']['COUNT_zero_results_scans'] / data['aggregate_metrics']['COUNT_scans']) * 100:.1f}%)")
    print(f"- Unique Projects Scanned: {format(data['aggregate_metrics']['COUNT_projects_scanned'], ',')}")
    return
    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"01-summary_of_scans.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Description','Value', '%'])
                writer.writerow(['Start Date',data['first_date']])
                writer.writerow(['End Date',data['last_date']])
                writer.writerow(['Days',total_days])
                writer.writerow(['Weeks',total_weeks])
                writer.writerow(['Scans Submitted',total_scan_count])
                writer.writerow(['Full Scans Submitted',full_scan_count,(full_scan_count / total_scan_count)])
                writer.writerow(['Incremental Scans Submitted',incremental_scan_count,(incremental_scan_count / total_scan_count)])
                writer.writerow(['No-Change Scans',no_scan_count,(no_scan_count / total_scan_count)])
                writer.writerow(['Scans with High Results',high_results__scan_count,(high_results__scan_count / total_scan_count)])
                writer.writerow(['Scans with Medium Results',medium_results__scan_count,(medium_results__scan_count / total_scan_count)])
                writer.writerow(['Scans with Low Results',low_results__scan_count,(low_results__scan_count / total_scan_count)])
                writer.writerow(['Scans with Informational Results',info_results__scan_count,(info_results__scan_count / total_scan_count)])
                writer.writerow(['Scans with Zero Results',zero_results__scan_count,(zero_results__scan_count / total_scan_count)])
                writer.writerow(['Unique Projects Scanned',len(data['vulnerability_stats_by_project'])])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Scan Metrics
    print("\nScan Metrics")
    print(f"- Avg LOC per Scan: {format(round(scan_loc__sum / total_scan_count), ',')}")
    print(f"- Max LOC per Scan:  {format(round(scan_loc__max), ',')}")
    print(f"- Avg Failed LOC per Scan: {format(round(scan_failed_loc__sum / yes_scan_count), ',')}")
    print(f"- Max Failed LOC per Scan:  {format(round(scan_failed_loc__max), ',')}")
    print(f"- Avg Daily LOC: {format(round(scan_loc__sum / total_scan_days), ',')}")
    print(f"- Max Daily LOC: {format(round(date_loc__max), ',')}")
    
    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"02-scan_metrics.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Description','Average', 'Max'])
                writer.writerow(['LOC per Scan',round(scan_loc__sum / total_scan_count),round(scan_loc__max)])
                writer.writerow(['Failed LOC per Scan',round(scan_failed_loc__sum / yes_scan_count),round(scan_failed_loc__max)])
                writer.writerow(['Daily LOC',round(scan_loc__sum / total_scan_days),round(date_loc__max)])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Scan Duration
    print("\nScan Duration")
    print(f"- Avg Total Scan Duration: {format_seconds_to_hms(total_scan_time__avg)}")
    print(f"- Max Total Scan Duration: {format_seconds_to_hms(total_scan_time__max)}")
    print(f"- Avg Engine Scan Duration: {format_seconds_to_hms(engine_scan_time__avg)}")
    print(f"- Max Engine Scan Duration: {format_seconds_to_hms(engine_scan_time__max)}")
    print(f"- Avg Queued Duration: {format_seconds_to_hms(queue_time__avg)}")
    print(f"- Max Queued Scan Duration: {format_seconds_to_hms(queue_time__max)}")
    print(f"- Avg Source Pulling Duration: {format_seconds_to_hms(source_pulling_time__avg)}")
    print(f"- Max Source Pulling Duration: {format_seconds_to_hms(source_pulling_time__max)}")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"03-scan_duration.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Description','Average', 'Max'])
                writer.writerow(['Total Scan Duration',format_seconds_to_hms(total_scan_time__avg),format_seconds_to_hms(total_scan_time__max)])
                writer.writerow(['Engine Scan Duration',format_seconds_to_hms(engine_scan_time__avg),format_seconds_to_hms(engine_scan_time__max)])
                writer.writerow(['Queued Duration',format_seconds_to_hms(queue_time__avg),format_seconds_to_hms(queue_time__max)])
                writer.writerow(['Source Pulling Duration',format_seconds_to_hms(source_pulling_time__avg),format_seconds_to_hms(source_pulling_time__max)])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Scan Results / Severity
    print("\nScan Results / Severity")
    print(f"- Average Total Results: {data['results']['total_vulns__avg']}")
    print(f"- Max Total Results: {data['results']['total_vulns__max']}")
    print(f"- Average High Results: {data['results']['high__avg']}")
    print(f"- Max High Results: {data['results']['high__max']}")
    print(f"- Average Medium Results: {data['results']['medium__avg']}")
    print(f"- Max Medium Results: {data['results']['medium__max']}")
    print(f"- Average Low Results: {data['results']['low__avg']}")
    print(f"- Max Low Results: {data['results']['low__max']}")
    print(f"- Average Informational Results: {data['results']['info__avg']}")
    print(f"- Max Informational Results: {data['results']['info__max']}")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"04-scan_results_severity.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Description','Average', 'Max'])
                writer.writerow(['Total',data['results']['total_vulns__avg'],data['results']['total_vulns__max']])
                writer.writerow(['High',data['results']['high__avg'],data['results']['high__max']])
                writer.writerow(['Medium',data['results']['medium__avg'],data['results']['medium__max']])
                writer.writerow(['Low',data['results']['low__avg'],data['results']['low__max']])
                writer.writerow(['Informational',data['results']['info__avg'],data['results']['info__max']])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")
    
    # Print Languages
    print("\nLanguages")
    for language_name, language_count in sorted(data['scanned_languages'].items(), key=lambda x: x[1], reverse=True):
        percentage = (language_count / total_scan_count) * 100
        print(f"- {language_name}: {format(language_count, ',')} ({percentage:.1f}%)")
    
    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"05-languages.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Language','%', 'Scans'])
                for language_name, language_count in sorted(data['scanned_languages'].items(), key=lambda x: x[1], reverse=True):
                    percentage = language_count / total_scan_count
                    writer.writerow([language_name,percentage,language_count])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Scan Submission Summary
    print("\nScan Submission Summary")
    print(f"- Average Scans Submitted per Week: {format(round(total_scan_count / total_weeks), ',')}")
    print(f"- Average Scans Submitted per Day: {format(round(total_scan_count / total_days), ',')}")
    print(f"- Average Scans Submitted per Week Day: {format(round(day_of_week_scan_totals['Weekday'] / (total_weeks * 5)), ',')}")
    print(f"- Average Scans Submitted per Weekend Day: {format(round(day_of_week_scan_totals['Weekend'] / (total_weeks * 2)), ',')}")
    print(f"- Max Daily Scans Submitted: {format(date_max_scan_count, ',')}")
    print(f"- Date of Max Scans: {date_max_scan_date}")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"06-scan_submissison_summary.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Description','Value'])
                writer.writerow(['Average Scans Submitted per Week',round(total_scan_count / total_weeks)])
                writer.writerow(['Average Scans Submitted per Day',round(total_scan_count / total_days)])
                writer.writerow(['Average Scans Submitted per Weekday',round(day_of_week_scan_totals['Weekday'] / (total_weeks * 5))])
                writer.writerow(['Average Scans Submitted per Weekend Day',round(day_of_week_scan_totals['Weekend'] / (total_weeks * 2))])
                writer.writerow(['Max Daily Scans Submitted',date_max_scan_count])
                writer.writerow(['Date of Max Scans',date_max_scan_date])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Day of Week Scan Average
    print("\nDay of Week Scan Average")
    for day_name, total_day_count in day_of_week_scan_totals.items():
        if day_name == "Weekday" or day_name == "Weekend":
            continue
        percentage = (total_day_count / total_scan_count) * 100
        print(f"- {day_name}: {format(round(total_day_count / total_weeks), ',')} ({percentage:.1f}%)")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"07-day_of_week_scan_average.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Day of Week', 'Scans', '%'])
                for day_name, total_day_count in day_of_week_scan_totals.items():
                    if day_name == "Weekday" or day_name == "Weekend":
                        continue
                    percentage = (total_day_count / total_scan_count)
                    writer.writerow([day_name,round(total_day_count / total_weeks),percentage])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Origins
    print("\nOrigins")
    for origin, origin_count in sorted(data['origins'].items(), key=lambda x: x[1], reverse=True):
        percentage = (origin_count / total_scan_count) * 100
        print(f"- {origin}: {format(origin_count, ',')} ({percentage:.1f}%)")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"08-origins.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Origin', 'Scans', '%'])
                for origin, origin_count in sorted(data['origins'].items(), key=lambda x: x[1], reverse=True):
                    percentage = (origin_count / total_scan_count)
                    writer.writerow([origin,origin_count,percentage])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Presets
    print("\nPresets")
    for preset_name, preset_count in sorted(data['preset_names'].items(), key=lambda x: x[1], reverse=True):
        percentage = (preset_count / total_scan_count) * 100
        print(f"- {preset_name}: {format(preset_count, ',')} ({percentage:.1f}%)")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"09-presets.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Preset', 'Scans', '%'])
                for preset_name, preset_count in sorted(data['preset_names'].items(), key=lambda x: x[1], reverse=True):
                    percentage = (preset_count / total_scan_count)
                    writer.writerow([preset_name,preset_count,percentage])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Print Scan Time Analysis
    print("\nScan Time Analysis")
    # Print the header of the table
    print(f"{'LOC Range':<12} {'Scans':<12} {'% Scans':<10} {'Avg Total':<18} {'Avg Src Pulling':<18} {'Avg Queue':<18} "
    f"{'Avg Engine':<18}")
    
    # Iterate through the size_bins dictionary to print the per-bin data
    for bin_key, bin_values in data['size_bins'].items():
        # Format times from seconds to HH:MM:SS
        source_pulling_time__avg = format_seconds_to_hms(bin_values['source_pulling_time__avg'])
        queue_time__avg = format_seconds_to_hms(bin_values['queue_time__avg'])
        engine_scan_time__avg = format_seconds_to_hms(bin_values['engine_scan_time__avg'])
        total_scan_time__avg = format_seconds_to_hms(bin_values['total_scan_time__avg'])
        
        # Print the formatted row for each bin
        print(f"{bin_key:<12} {bin_values['yes_scan_count'] + bin_values['no_scan_count']:<12,} "
        f"{(math.ceil((10000 * (bin_values['yes_scan_count'] + bin_values['no_scan_count']) / total_scan_count)) / 100):<11.2f}"
        f"{total_scan_time__avg:<18} {source_pulling_time__avg:<18} {queue_time__avg:<18} {engine_scan_time__avg:<18}")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"10-scan_time_analysis.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['LOC Range','Scans','% Scans','Avg Total Time','Avg Source Pulling Time','Avg Queue Time','Avg Engine Scan Time'])
                
                # Iterate through the size_bins dictionary to print the per-bin data
                for bin_key, bin_values in data['size_bins'].items():
                    # Format times from seconds to HH:MM:SS
                    source_pulling_time__avg = format_seconds_to_hms(bin_values['source_pulling_time__avg'])
                    queue_time__avg = format_seconds_to_hms(bin_values['queue_time__avg'])
                    engine_scan_time__avg = format_seconds_to_hms(bin_values['engine_scan_time__avg'])
                    total_scan_time__avg = format_seconds_to_hms(bin_values['total_scan_time__avg'])
                    writer.writerow([bin_key,bin_values['yes_scan_count'] + bin_values['no_scan_count'],
                        math.ceil((10000 * (bin_values['yes_scan_count'] + bin_values['no_scan_count']) / total_scan_count)) / 10000,
                        total_scan_time__avg,source_pulling_time__avg,queue_time__avg,engine_scan_time__avg])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")
    
    # Print Concurrency Summary with unique and sorted dates
    print("\nConcurrency Summary")
    print(f"- Overall Peak Actual Concurrency: {overall_max_actual} concurrent scans on {', '.join(map(str, overall_max_actual_dates))}")
    print(f"- Overall Peak Optimal Concurrency: {overall_max_optimal} concurrent scans on {', '.join(map(str, overall_max_optimal_dates))}")

    # Create output file, if required
    if csv_config['enabled']:
        try:
            filename = os.path.join(csv_config['csv_dir'], f"11-concurrency_analysis.csv")
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['Date', 'Max Actual', 'Max Optimal'])
                for date, maxima in sorted(daily_maxima.items()):
                    writer.writerow([date, maxima['actual'], maxima['optimal']])
        except IOError as e:
            print(f"IOError when writing to file: {e}")
        except Exception as e:
            print(f"Unexpected error when creating/writing to the CSV file: {e}")

    # Create other output files for data that we don't print to the summary, if required
    
    if csv_config['enabled']:
        write_to_csv(["Date","Scans"],sorted(daily_scan_counts.items()), os.path.join(csv_config['csv_dir'], f"12-scans_by_date.csv"))
        write_to_csv(["Week","Scans"],sorted(weekly_scan_counts.items()), os.path.join(csv_config['csv_dir'], f"13-scans_by_week.csv"))

    if excel_config['enabled']:
        write_to_excel(sorted(daily_scan_counts.items()), "AW", 4)
        write_to_excel(sorted(weekly_scan_counts.items()), "AZ", 4)

    print("")



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process scans and output CSV files if requested.")
    parser.add_argument("input_file", type=str, help="The JSON file containing scan data.")
    parser.add_argument("--csv", action="store_true", help="Generate CSV output files.")
    parser.add_argument("--full_data", action="store_true", help="Generate CSV output of complete scan data.")
    parser.add_argument("--name", type=str, default="", help="Optional name for the output directory")
    parser.add_argument("--excel", type=str, default="", help="Export EHC data directly to the specified Excel workbook")

    args = parser.parse_args()
    input_file = args.input_file
    output_name = args.name if args.name else os.path.splitext(os.path.basename(input_file))[0]
    excel_filename = args.excel if args.excel else None
    
    # Define the output directory using the optional name if provided
    csv_dir = os.path.join(os.getcwd(), f"ehc_output_{output_name}_{datetime.now().strftime('%Y%m%d-%H%M%S')}")

    # Initialize structures to hold output configs
    full_csv = {
        'enabled': True if args.full_data else False,
        'csv_dir': csv_dir,
        'field_names': []
    }
    csv_config = {
        'enabled': True if args.csv else False,
        'csv_dir': csv_dir
    }
    excel_config = {
        'enabled': True if args.excel else False,
        'filename': excel_filename
    }
    
    # If we're exporting to Excel...
    if excel_config['enabled']:
        # Make sure we have the required libraries
        if not xl_available:
            parser.error("--excel requires pandas and openpyxl libraries: 'pip install pandas openpyxl'")
        # Make sure the file exists
        if not os.path.isfile(excel_filename):
            print(f"Error: The file '{excel_filename}' does not exist.")
            exit(1)
        try:
            # Open the workbook
            workbook = load_workbook(excel_filename)
        except InvalidFileException:
            print(f"Error: The file '{excel_filename}' is not a valid Excel file or is corrupted.")
            exit(1)
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            exit(1)

    # If we are creating any CSV files, create the output directory i
    if args.full_data or args.csv:
        try:
            # Attempt to create the directory
            os.makedirs(csv_dir, exist_ok=True)
        except PermissionError as e:
            print(f"Permission Error: {e}")
            exit(1)
        except Exception as e:
            print(f"Error creating directory: {e}")
            exit(1)

    full_csv['field_names'], scans = ingest_file(input_file)

    processed_data = process_scans(scans, full_csv)

    output_analysis(processed_data, csv_config, excel_config)

    # If we exported to Excel, save the workbook
    if(excel_config['enabled']):
        workbook.save(excel_filename)